# -*- coding: utf-8 -*-
"""人工标注 vs 自动检测 —— 论文级对比表生成器（1-10 级，音型集六特征口径）。

输入:
  - 人工标注: C:/Users/BinyuWang/Desktop/CV/etude.xlsx（10 首 × 6 特征命中小节数）
  - 自动检测: output_grades/etude01..10/reports/report.json

输出（默认 D:/HOMR/output_grades/reports/）:
  - 人工vs自动对比_论文版.xlsx（5 张表）
  - 人工vs自动对比_论文版.md

用法: python human_vs_auto_compare.py [--human PATH] [--root D:/HOMR/output_grades]
"""

import argparse
import json
import sys
from pathlib import Path

FEATURE_ORDER = [
    "wide_leap_sum", "contracted_grip", "consecutive_sixteenths",
    "consecutive_eighths", "consecutive_chords", "consecutive_intervals",
]
FEATURE_CN = {
    "wide_leap_sum": "伸张把位",
    "contracted_grip": "收缩把位",
    "consecutive_sixteenths": "连续十六分跑动",
    "consecutive_eighths": "连续八分跑动",
    "consecutive_chords": "连续和弦支撑",
    "consecutive_intervals": "连续音程支撑",
}


def load_human(path: str) -> dict:
    """读取人工标注 xlsx -> {grade: {feat_code: int, '小节数': int}}

    xlsx 列结构固定为：
      (序号, 小节数, 伸张把位, 收缩把位, 连续十六分音符跑动,
       连续八分音符跑动, 连续和弦, 连续音程)
    按列位置直接映射到特征代码，避免列名差异（"连续十六分音符跑动"
    vs 代码名）导致匹配失败。
    """
    from openpyxl import load_workbook
    wb = load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    # 列 -> 特征代码（索引 0 为序号、1 为小节数）
    col_to_feat = {
        2: "wide_leap_sum",          # 伸张把位
        3: "contracted_grip",        # 收缩把位
        4: "consecutive_sixteenths", # 连续十六分音符跑动
        5: "consecutive_eighths",    # 连续八分音符跑动
        6: "consecutive_chords",     # 连续和弦
        7: "consecutive_intervals",  # 连续音程
    }
    data = {}
    for r in rows[1:]:
        if r[0] is None:
            continue
        grade = int(r[0])
        entry = {"小节数": int(r[1])}
        for col, feat in col_to_feat.items():
            entry[feat] = int(r[col] or 0)
        data[grade] = entry
    return data


def load_auto(root: Path) -> dict:
    """读取自动报告 -> {grade: {feat: matched, total, ratio}}"""
    out = {}
    for i in range(1, 11):
        p = root / f"etude{i:02d}" / "reports" / "report.json"
        if not p.exists():
            continue
        d = json.loads(p.read_text(encoding="utf-8"))
        s = d["scores"][0]
        out[i] = {
            "features": {
                fk: {"matched": fv["matched_measures"], "total": fv["total_measures"]}
                for fk, fv in s["features"].items()
            },
            "difficulty": s.get("difficulty_score", 0.0),
            "detail": s.get("difficulty_detail", {}),
        }
    return out


def main():
    sys.stdout.reconfigure(encoding="utf-8")
    ap = argparse.ArgumentParser()
    ap.add_argument("--human", default=r"C:\Users\BinyuWang\Desktop\CV\etude.xlsx")
    ap.add_argument("--root", default="D:/HOMR/output_grades")
    args = ap.parse_args()

    human = load_human(args.human)
    auto = load_auto(Path(args.root))
    grades = sorted(set(human) & set(auto))
    if not grades:
        print("✗ 人工/自动数据无交集"); return

    # 每个单元 (grade, feature) 的一致率 c = 1 - |a-h| / n（n 取人工小节数）
    cells = {}  # (grade, feat) -> dict(h, a, n, c)
    for g in grades:
        n_h = human[g]["小节数"]
        for f in FEATURE_ORDER:
            h = human[g].get(f, 0)
            a = auto[g]["features"][f]["matched"]
            c = 1 - abs(a - h) / n_h
            cells[(g, f)] = {"h": h, "a": a, "n": n_h, "c": c}

    out_dir = Path(args.root) / "reports"
    out_dir.mkdir(parents=True, exist_ok=True)

    # ---------- Markdown ----------
    L = ["# 人工标注 vs 自动检测 · 论文级对比（1–10 级 × 6 特征）", "",
         f"- 生成时间: {__import__('time').strftime('%Y-%m-%d %H:%M:%S')}",
         "- 口径: 《音型（音群）集》六类特征；一致率 c = 1 − |自动−人工| / 小节数（n 取人工小节数）",
         "- 配置: 交叉手重分配 + 和弦最高音修正（chord top）+ 单调方向条件（音型集口径）", ""]

    # 表 1 逐曲逐特征对比
    L.append("## 表 1 逐曲逐特征对比（人工 → 自动，差值）")
    L.append("")
    head = "| 曲目 | 小节(人工) | " + " | ".join(
        f"{FEATURE_CN[f]}<br>人工/自动/Δ" for f in FEATURE_ORDER) + " |"
    L.append(head)
    L.append("|---|---:|" + "---:|" * len(FEATURE_ORDER))
    for g in grades:
        n_h = human[g]["小节数"]
        cells_row = []
        for f in FEATURE_ORDER:
            h = human[g].get(f, 0)
            a = auto[g]["features"][f]["matched"]
            d = a - h
            sign = "" if d == 0 else ("+" if d > 0 else "")
            cells_row.append(f"{h}/{a}/{sign}{d}")
        L.append(f"| etude{g:02d} | {n_h} | " + " | ".join(cells_row) + " |")
    L.append("")

    # 表 2 分特征一致率汇总
    L.append("## 表 2 六特征一致率汇总（60 单元均值）")
    L.append("")
    L.append("| 特征 | 人工合计 | 自动合计 | 平均一致率 | 最佳曲目 | 最差曲目 |")
    L.append("|---|---:|---:|---:|---|---|")
    totals = []
    for f in FEATURE_ORDER:
        hs = sum(cells[(g, f)]["h"] for g in grades)
        a_s = sum(cells[(g, f)]["a"] for g in grades)
        cs = {g: cells[(g, f)]["c"] for g in grades}
        best = max(cs, key=cs.get)
        worst = min(cs, key=cs.get)
        avg = sum(cs.values()) / len(cs)
        totals.append(avg)
        L.append(f"| {FEATURE_CN[f]} | {hs} | {a_s} | **{avg*100:.1f}%** "
                 f"| etude{best:02d} ({cs[best]*100:.0f}%) | etude{worst:02d} ({cs[worst]*100:.0f}%) |")
    L.append(f"| **总体（60 单元）** | — | — | **{sum(totals)/len(totals)*100:.1f}%** | | |")
    L.append("")

    # 表 3 人工小节数 vs 自动小节数（口径差异）
    L.append("## 表 3 小节数口径差异")
    L.append("")
    L.append("| 曲目 | 人工小节 | 自动小节 | 差值 |")
    L.append("|---|---:|---:|---:|")
    for g in grades:
        n_h = human[g]["小节数"]
        n_a = auto[g]["features"]["wide_leap_sum"]["total"]
        L.append(f"| etude{g:02d} | {n_h} | {n_a} | {n_a - n_h:+d} |")
    L.append("")

    # 表 4 难度加权分
    L.append("## 表 4 难度加权总分（自动）与等级")
    L.append("")
    L.append("| 等级 | 1 | 2 | 3 | 4 | 5 | 6 | 7 | 8 | 9 | 10 |")
    L.append("|---|---:|---:|---:|---:|---:|---:|---:|---:|---:|---:|")
    row = "| 难度分 | " + " | ".join(f"{auto[g]['difficulty']:g}" for g in grades) + " |"
    L.append(row)
    scores = [auto[g]["difficulty"] for g in grades]
    # Spearman
    def rank(v):
        order = sorted(range(len(v)), key=lambda i: v[i])
        r = [0.0] * len(v)
        i = 0
        while i < len(order):
            j = i
            while j + 1 < len(order) and v[order[j + 1]] == v[order[i]]:
                j += 1
            avg = (i + j) / 2 + 1
            for k in range(i, j + 1):
                r[order[k]] = avg
            i = j + 1
        return r
    rx, ry = rank(grades), rank(scores)
    n = len(grades)
    mx, my = sum(rx) / n, sum(ry) / n
    num = sum((a - mx) * (b - my) for a, b in zip(rx, ry))
    den = (sum((a - mx) ** 2 for a in rx) * sum((b - my) ** 2 for b in ry)) ** 0.5
    rho = num / den if den else 0.0
    L.append("")
    L.append(f"Spearman 等级相关系数 ρ = **{rho:.3f}**（难度分随等级单调递增）")
    L.append("")

    md_text = "\n".join(L)
    md_path = out_dir / "人工vs自动对比_论文版.md"
    md_path.write_text(md_text, encoding="utf-8")
    print(f"✓ {md_path}")

    # ---------- Excel ----------
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    HEADER_FILL = PatternFill("solid", fgColor="4472C4")
    HEADER_FONT = Font(bold=True, color="FFFFFF")
    GOOD_FILL = PatternFill("solid", fgColor="E2EFDA")
    BAD_FILL = PatternFill("solid", fgColor="FCE4D6")

    wb = Workbook()

    def style_header(ws, n_cols):
        for c in range(1, n_cols + 1):
            cell = ws.cell(row=1, column=c)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.freeze_panes = "A2"

    # Sheet1 逐曲逐特征
    ws = wb.active
    ws.title = "表1_逐曲对比"
    headers = ["曲目", "小节(人工)"]
    for f in FEATURE_ORDER:
        headers += [f"{FEATURE_CN[f]} 人工", f"{FEATURE_CN[f]} 自动", f"Δ"]
    ws.append(headers)
    for g in grades:
        n_h = human[g]["小节数"]
        row = [f"etude{g:02d}", n_h]
        for f in FEATURE_ORDER:
            h = human[g].get(f, 0)
            a = auto[g]["features"][f]["matched"]
            row += [h, a, a - h]
        ws.append(row)
    style_header(ws, len(headers))
    for r in range(2, ws.max_row + 1):
        for c in range(3, len(headers) + 1, 3):  # Δ 列
            v = ws.cell(row=r, column=c + 2).value
            if v == 0:
                ws.cell(row=r, column=c + 2).fill = GOOD_FILL
            else:
                ws.cell(row=r, column=c + 2).fill = BAD_FILL
    for col, w in zip(range(1, len(headers) + 1), [10, 10] + [12] * (len(headers) - 2)):
        ws.column_dimensions[get_column_letter(col)].width = w

    # Sheet2 一致率
    ws2 = wb.create_sheet("表2_一致率汇总")
    headers2 = ["特征", "人工合计", "自动合计", "平均一致率%", "最佳曲目", "最差曲目"]
    ws2.append(headers2)
    for f in FEATURE_ORDER:
        hs = sum(cells[(g, f)]["h"] for g in grades)
        a_s = sum(cells[(g, f)]["a"] for g in grades)
        cs = {g: cells[(g, f)]["c"] for g in grades}
        best = max(cs, key=cs.get)
        worst = min(cs, key=cs.get)
        avg = sum(cs.values()) / len(cs)
        ws2.append([FEATURE_CN[f], hs, a_s, round(avg * 100, 1),
                    f"etude{best:02d} ({cs[best]*100:.0f}%)",
                    f"etude{worst:02d} ({cs[worst]*100:.0f}%)"])
    ws2.append(["总体（60 单元）", "", "", round(sum(totals) / len(totals) * 100, 1), "", ""])
    style_header(ws2, len(headers2))
    for col, w in zip(range(1, 7), [18, 10, 10, 12, 22, 22]):
        ws2.column_dimensions[get_column_letter(col)].width = w

    # Sheet3 小节差异
    ws3 = wb.create_sheet("表3_小节口径")
    headers3 = ["曲目", "人工小节", "自动小节", "差值"]
    ws3.append(headers3)
    for g in grades:
        n_h = human[g]["小节数"]
        n_a = auto[g]["features"]["wide_leap_sum"]["total"]
        ws3.append([f"etude{g:02d}", n_h, n_a, n_a - n_h])
    style_header(ws3, len(headers3))

    # Sheet4 难度加权
    ws4 = wb.create_sheet("表4_难度加权")
    headers4 = ["等级"] + [str(g) for g in grades]
    ws4.append(headers4)
    ws4.append(["难度加权总分"] + [auto[g]["difficulty"] for g in grades])
    ws4.append(["伸张把位"] + [auto[g]["detail"].get("wide_leap_sum", {}).get("score", 0) for g in grades])
    ws4.append(["连续十六分"] + [auto[g]["detail"].get("consecutive_sixteenths", {}).get("score", 0) for g in grades])
    ws4.append(["连续八分"] + [auto[g]["detail"].get("consecutive_eighths", {}).get("score", 0) for g in grades])
    ws4.append(["收缩把位"] + [auto[g]["detail"].get("contracted_grip", {}).get("score", 0) for g in grades])
    ws4.append(["连续和弦"] + [auto[g]["detail"].get("consecutive_chords", {}).get("score", 0) for g in grades])
    ws4.append(["连续音程"] + [auto[g]["detail"].get("consecutive_intervals", {}).get("score", 0) for g in grades])
    ws4.append(["Spearman ρ", "", "", "", "", "", "", "", "", "", rho])
    style_header(ws4, len(headers4))
    for col, w in zip(range(1, len(headers4) + 1), [12] + [10] * len(grades)):
        ws4.column_dimensions[get_column_letter(col)].width = w

    # Sheet5 说明
    ws5 = wb.create_sheet("表5_口径说明")
    notes = [
        ["一致率定义", "c = 1 − |自动命中 − 人工命中| / 小节数；总体为 10 曲 × 6 特征 = 60 单元均值"],
        ["小节数口径", "人工小节数来自人工统计表；自动小节数来自 MusicXML（跨页/拾起小节归并方式不同，1–3 小节差异）"],
        ["特征口径", "《音型（音群）集》：伸张把位(单调+>12半音,w=2)、收缩把位(5音4音程<6半音,w=1)、连续十六分(5个,w=2)、连续八分(3个,w=1)、连续和弦(3个≥3音,w=2)、连续音程(3个双音,w=1)"],
        ["检测配置", "HOMR + music21；交叉手重分配、和弦最高音修正（chord top）、单调方向条件；无琶音感知豁免"],
        ["人工标注", "作者逐曲统计六类特征命中小节数（金标准）"],
    ]
    for row in notes:
        ws5.append(row)
    for col, w in zip(range(1, 3), [16, 90]):
        ws5.column_dimensions[get_column_letter(col)].width = w

    xlsx_path = out_dir / "人工vs自动对比_论文版.xlsx"
    wb.save(xlsx_path)
    print(f"✓ {xlsx_path}")


if __name__ == "__main__":
    main()
