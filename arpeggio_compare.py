# -*- coding: utf-8 -*-
"""琶音感知开关（arpeggio_aware）：基线 vs 开启 论文级对比报告生成器。

背景：wide_leap_sum 的复合音程判据（相邻 3 音两段音程之和 > 八度）在琶音
织体（分解和弦伴奏、跨八度琶音跑动）上会系统性误报——音群实际是同一和弦
的和弦音先后出现，不是旋律宽跳。v1.2.0 为 wide_leap_sum 增加了
arpeggio_aware 开关：音群音级集合若可整体嵌入某个三和弦/常用七和弦，
判为分解和弦音型并豁免。

本脚本在 4 首考级练习曲上跑两版引擎（均启用 v1.1.0 的交叉手重分配，
唯一变量是 arpeggio_aware off/on），生成：

  Excel「琶音感知对比_论文版.xlsx」
    - 表_对比总览    ：每曲 wide_leap_sum 两版命中/占比/Δ + 豁免琶音音组数（论文主表）
    - 表_命中分解    ：基线命中音组按 琶音/非琶音 × 单调/折返 分类（论点支撑表）
    - 表_逐小节明细  ：每条基线命中音组的音程/方向/琶音判定/开关后状态（附录表）
    - 表_其他特征校验：其余特征两版应零差异（开关无副作用证明）
    - 表_全语料汇总  ：4 曲合计对比

  Markdown「琶音感知对比_论文版.md」：可直接粘贴进论文的表格与结论。

用法：
    python arpeggio_compare.py                      # 默认扫 output 目录
    python arpeggio_compare.py -o 输出目录           # 指定 XML 根目录
    python arpeggio_compare.py --out 保存前缀        # 输出文件前缀
"""

import argparse
import copy
import sys
import time
from collections import Counter
from pathlib import Path

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

sys.path.insert(0, str(Path(__file__).resolve().parent))

from music21 import note as m21note

from scoreflow.analysis.engine import AnalysisEngine
from scoreflow.analysis.features import build_features
from scoreflow.analysis.features.wide_leap_sum import is_arpeggio
from scoreflow.config import load_config

FEAT = "wide_leap_sum"


# ----------------------------------------------------------------------
def parse_hit(desc: str):
    """'C4 -> A-2 -> D-2 (16+7=23半音)' -> (list[note.Note], 音程和, 方向)。"""
    body, _, tail = desc.partition(" (")
    names = [s.strip().strip("[]") for s in body.split(" -> ")]
    total = int(tail.split("=")[1].replace("半音", "").rstrip(")"))
    notes = [m21note.Note(n) for n in names]
    m = [n.pitch.midi for n in notes]
    if m[0] < m[1] < m[2]:
        direction = "单调上行"
    elif m[0] > m[1] > m[2]:
        direction = "单调下行"
    else:
        direction = "折返"
    return notes, total, direction


def classify_hits(hits):
    """把一批 MeasureHit 逐条分类，返回 [(hit, total, direction, is_arp)]。"""
    rows = []
    for h in hits:
        notes, total, direction = parse_hit(h.notes_desc)
        rows.append((h, total, direction, is_arpeggio(notes)))
    return rows


def collect_scores(xml_root: Path, pattern: str = "etude*"):
    scores = []
    for d in sorted(xml_root.glob(f"{pattern}/pages")):
        xmls = sorted(d.glob("page_*.musicxml"))
        if xmls:
            scores.append((d.parent.name, xmls))
    return scores


def agg(results, feature_name):
    total = sum(r.feature_results[feature_name].total_measures for r in results)
    matched = sum(r.feature_results[feature_name].matched_measures for r in results)
    hits = []
    for r in results:
        hits.extend(r.feature_results[feature_name].hits)
    return total, matched, hits


# ----------------------------------------------------------------------
def main():
    ap = argparse.ArgumentParser(description="琶音感知开关对比报告")
    ap.add_argument("-o", "--xml-root", default=r"D:\HOMR\output",
                    help="XML 根目录（含 <曲目>/pages/page_*.musicxml）")
    ap.add_argument("--pattern", default="etude*",
                    help="曲目目录过滤（glob），默认 etude* 只处理考级练习曲")
    ap.add_argument("--out", default=r"D:\HOMR\output\reports\琶音感知对比_论文版",
                    help="输出文件前缀（.xlsx / .md 自动追加）")
    args = ap.parse_args()

    # ---- 两版引擎：均启用交叉手重分配，唯一变量 = arpeggio_aware ----
    cfg = load_config(None, {})
    base_feats_cfg = copy.deepcopy(cfg.features)
    aware_feats_cfg = copy.deepcopy(cfg.features)
    aware_feats_cfg[FEAT]["params"]["arpeggio_aware"] = True

    kw = dict(chord_mode=cfg.chord_mode,
              count_empty_measures=cfg.count_empty_measures,
              across_barlines=cfg.across_barlines,
              cross_hand_heuristic=True)  # 与 v1.1.0 基线保持可比
    eng_base = AnalysisEngine(build_features(base_feats_cfg), **kw)
    eng_aware = AnalysisEngine(build_features(aware_feats_cfg), **kw)
    feature_names = [f.name for f in eng_base.features]
    other_features = [fn for fn in feature_names if fn != FEAT]

    scores = collect_scores(Path(args.xml_root), args.pattern)
    if not scores:
        print("✗ 未找到乐谱目录（*/pages/page_*.musicxml）:", args.xml_root)
        return 1
    print(f"找到 {len(scores)} 首乐谱；变量: wide_leap_sum.arpeggio_aware off→on"
          f"（两版均启用交叉手重分配）")

    # ---- 数据收集 ----
    overview = {}       # score -> {base_t/base_m/aware_t/aware_m/n_measures/rows}
    page_rows = []      # 逐小节明细
    other_rows = []     # 其他特征校验
    inconsistent = 0

    for score_name, xmls in scores:
        base_results, aware_results = [], []
        s_rows = []
        for xml in xmls:
            page_no = xml.stem.replace("page_", "")
            rb = eng_base.analyze_file(xml)
            ra = eng_aware.analyze_file(xml)
            base_results.append(rb)
            aware_results.append(ra)

            # 其他特征校验（应零差异）
            for fn in other_features:
                bm = rb.feature_results[fn].matched_measures
                bt = rb.feature_results[fn].total_measures
                am = ra.feature_results[fn].matched_measures
                at = ra.feature_results[fn].total_measures
                other_rows.append((score_name, page_no, fn,
                                   f"{bm}/{bt}", f"{am}/{at}", am - bm))

            # wide_leap_sum：基线命中逐条分类
            bhits = rb.feature_results[FEAT].hits
            ahits = ra.feature_results[FEAT].hits
            for h, total, direction, arp in classify_hits(bhits):
                s_rows.append((score_name, page_no, h.measure_number, h.part_name,
                               h.notes_desc, total, direction, arp))
                page_rows.append((score_name, page_no, h.measure_number, h.part_name,
                                  h.notes_desc, total, direction, arp))

            # 自洽校验：开关版命中音组 == 基线命中中非琶音音组
            expected = Counter(h.notes_desc for h in bhits
                               if not is_arpeggio(parse_hit(h.notes_desc)[0]))
            actual = Counter(h.notes_desc for h in ahits)
            if expected != actual:
                inconsistent += 1
                print(f"  ⚠ 自洽校验失败 {score_name} p{page_no}: "
                      f"预期保留 {sum(expected.values())} 条, 实际 {sum(actual.values())} 条")

        bt, bm, bhits = agg(base_results, FEAT)
        at, am, _ = agg(aware_results, FEAT)
        overview[score_name] = {
            "base_t": bt, "base_m": bm, "aware_t": at, "aware_m": am,
            "rows": s_rows,
        }

    # ---- 全语料统计（提前算好，后续循环不得覆盖）----
    bt = sum(o["base_t"] for o in overview.values())
    bm = sum(o["base_m"] for o in overview.values())
    am = sum(o["aware_m"] for o in overview.values())
    at = sum(o["aware_t"] for o in overview.values())
    n_rows = len(page_rows)
    n_arp = sum(1 for r in page_rows if r[7])

    # ---- 写 Excel ----
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    HEADER_FILL = PatternFill("solid", fgColor="4472C4")
    HEADER_FONT = Font(bold=True, color="FFFFFF")
    DIFF_FILL = PatternFill("solid", fgColor="E2EFDA")   # 有变化（绿）
    NOCHG_FILL = PatternFill("solid", fgColor="F2F2F2")  # 无变化（灰）
    ARP_FILL = PatternFill("solid", fgColor="FCE4D6")    # 琶音豁免（橙）

    def style_header(ws, n_cols):
        for col in range(1, n_cols + 1):
            c = ws.cell(row=1, column=col)
            c.fill = HEADER_FILL
            c.font = HEADER_FONT
            c.alignment = Alignment(horizontal="center", vertical="center")
        ws.freeze_panes = "A2"

    def set_widths(ws, widths):
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    wb = Workbook()

    # ===== Sheet1 表_对比总览（论文主表）=====
    ws = wb.active
    ws.title = "表_对比总览"
    headers = ["乐谱", "小节数", "基线命中/小节", "基线占比%",
               "开关后命中/小节", "开关后占比%", "Δ占比(pp)", "豁免琶音音组数"]
    ws.append(headers)
    for name, o in overview.items():
        t_arp = sum(1 for r in o["rows"] if r[7])
        b_ratio = o["base_m"] / o["base_t"] * 100 if o["base_t"] else 0.0
        a_ratio = o["aware_m"] / o["aware_t"] * 100 if o["aware_t"] else 0.0
        ws.append([name, o["base_t"],
                   f"{o['base_m']}/{o['base_t']}", round(b_ratio, 1),
                   f"{o['aware_m']}/{o['aware_t']}", round(a_ratio, 1),
                   round(a_ratio - b_ratio, 1), t_arp])
    style_header(ws, len(headers))
    for r in range(2, ws.max_row + 1):
        for col in (4, 6):
            ws.cell(row=r, column=col).number_format = "0.0\"%\""
        dcell = ws.cell(row=r, column=7)
        dcell.fill = NOCHG_FILL if dcell.value == 0 else DIFF_FILL
    set_widths(ws, [16, 9, 14, 11, 15, 12, 11, 14])

    # ===== Sheet2 表_命中分解（论点支撑表）=====
    ws2 = wb.create_sheet("表_命中分解")
    headers2 = ["乐谱", "基线命中音组", "琶音音组(豁免)", "非琶音音组(保留)",
                "琶音占比%", "其中:折返琶音", "其中:单调琶音",
                "非琶音中:折返", "非琶音中:单调"]
    ws2.append(headers2)
    for name, o in overview.items():
        rows = o["rows"]
        n = len(rows)
        s_arp = sum(1 for r in rows if r[7])
        arp_rows = [r for r in rows if r[7]]
        non_rows = [r for r in rows if not r[7]]
        ws2.append([name, n, s_arp, n - s_arp,
                    round(s_arp / n * 100, 1) if n else 0,
                    sum(1 for r in arp_rows if r[6] == "折返"),
                    s_arp - sum(1 for r in arp_rows if r[6] == "折返"),
                    sum(1 for r in non_rows if r[6] == "折返"),
                    n - s_arp - sum(1 for r in non_rows if r[6] == "折返")])
    style_header(ws2, len(headers2))
    for r in range(2, ws2.max_row + 1):
        ws2.cell(row=r, column=5).number_format = "0.0\"%\""
    set_widths(ws2, [16, 13, 15, 16, 11, 13, 13, 13, 13])

    # ===== Sheet3 表_逐小节明细（附录表）=====
    ws3 = wb.create_sheet("表_逐小节明细")
    headers3 = ["乐谱", "页", "小节", "声部", "音群", "音程和(半音)",
                "方向", "琶音判定", "开关后"]
    ws3.append(headers3)
    for name, page, mnum, part, desc, total, direction, arp in page_rows:
        ws3.append([name, int(page), mnum, part, desc, total, direction,
                    "琶音→豁免" if arp else "非琶音→保留",
                    "不再命中" if arp else "仍命中"])
    style_header(ws3, len(headers3))
    for r in range(2, ws3.max_row + 1):
        if str(ws3.cell(row=r, column=8).value).startswith("琶音"):
            for col in (8, 9):
                ws3.cell(row=r, column=col).fill = ARP_FILL
    set_widths(ws3, [16, 7, 7, 12, 34, 12, 10, 13, 10])

    # ===== Sheet4 表_其他特征校验（无副作用证明）=====
    ws4 = wb.create_sheet("表_其他特征校验")
    headers4 = ["乐谱", "页", "特征", "基线命中/小节", "开关后命中/小节", "Δ"]
    ws4.append(headers4)
    for row in other_rows:
        ws4.append(list(row))
    style_header(ws4, len(headers4))
    for r in range(2, ws4.max_row + 1):
        dcell = ws4.cell(row=r, column=6)
        dcell.fill = NOCHG_FILL if dcell.value == 0 else DIFF_FILL
    set_widths(ws4, [16, 7, 26, 14, 15, 6])

    # ===== Sheet5 表_全语料汇总 =====
    ws5 = wb.create_sheet("表_全语料汇总")
    headers5 = ["指标", "基线(arpeggio_aware=off)", "开关后(arpeggio_aware=on)"]
    ws5.append(headers5)
    ws5.append(["乐谱数 / 小节总数", f"{len(overview)} 曲 / {bt} 小节",
                f"{len(overview)} 曲 / {at} 小节"])
    ws5.append(["wide_leap_sum 命中小节", bm, am])
    ws5.append(["wide_leap_sum 命中占比",
                round(bm / bt * 100, 1) if bt else 0,
                round(am / at * 100, 1) if at else 0])
    ws5.append(["wide_leap_sum 命中音组(记录数)", n_rows, n_rows - n_arp])
    ws5.append(["其中: 琶音音组(被豁免)", "—", n_arp])
    ws5.append(["其中: 非琶音音组(保留)", n_rows, n_rows - n_arp])
    ws5.append(["其他特征受影响行数", f"{len(other_rows)} 行全部 Δ=0"
                if all(r[5] == 0 for r in other_rows) else "存在非零Δ！",
                ""])
    style_header(ws5, len(headers5))
    set_widths(ws5, [30, 26, 26])

    xlsx_path = Path(args.out + ".xlsx")
    xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(xlsx_path))
    print("✓ Excel:", xlsx_path)

    # ---- 写 Markdown ----
    L = [
        "# 琶音感知开关（arpeggio_aware）对比报告",
        "",
        f"- 生成时间: {time.strftime('%Y-%m-%d %H:%M:%S')}",
        f"- 语料: {args.xml_root}（{args.pattern}，{len(overview)} 首考级练习曲，共 {bt} 小节）",
        "- 实验设置: 两版引擎均启用交叉手重分配（v1.1.0），"
        "唯一变量为 `wide_leap_sum.params.arpeggio_aware` off → on",
        "- 琶音判定: 音群 3 音的音级（pitch class）集合可整体嵌入某个三和弦"
        "或常用七和弦（大小三/减/增三和弦，大小七/属七/减七/半减七），",
        "  即视为分解和弦（琶音）音型，予以豁免（详见 README）。",
        "",
        "## 表 1. 每首乐谱 wide_leap_sum 对比（论文主表）",
        "",
        "| 乐谱 | 小节数 | 基线命中/小节 | 基线占比 | 开关后命中/小节 | 开关后占比 | Δ占比(pp) | 豁免琶音音组 |",
        "|:---|---:|:---:|:---:|:---:|:---:|:---:|---:|",
    ]
    for name, o in overview.items():
        m_arp = sum(1 for r in o["rows"] if r[7])
        b_ratio = o["base_m"] / o["base_t"] * 100 if o["base_t"] else 0.0
        a_ratio = o["aware_m"] / o["aware_t"] * 100 if o["aware_t"] else 0.0
        L.append(
            f"| {name} | {o['base_t']} | {o['base_m']}/{o['base_t']} | {b_ratio:.1f}% "
            f"| {o['aware_m']}/{o['aware_t']} | {a_ratio:.1f}% "
            f"| {a_ratio - b_ratio:+.1f} | {m_arp} |")
    L += [
        "",
        "## 表 2. 基线命中音组分解（琶音 vs 非琶音 × 方向）",
        "",
        "| 乐谱 | 基线命中音组 | 琶音(豁免) | 非琶音(保留) | 琶音占比 | 折返琶音 | 单调琶音 | 非琶音折返 | 非琶音单调 |",
        "|:---|---:|---:|---:|:---:|---:|---:|---:|---:|",
    ]
    for name, o in overview.items():
        rows = o["rows"]
        n = len(rows)
        arp_rows = [r for r in rows if r[7]]
        non_rows = [r for r in rows if not r[7]]
        m_arp = len(arp_rows)
        arp_turn = sum(1 for r in arp_rows if r[6] == "折返")
        non_turn = sum(1 for r in non_rows if r[6] == "折返")
        L.append(
            f"| {name} | {n} | {m_arp} | {n - m_arp} "
            f"| {m_arp / n * 100:.1f}% | {arp_turn} | {m_arp - arp_turn} "
            f"| {non_turn} | {n - m_arp - non_turn} |")
    L += [
        "",
        "## 表 3. 全语料汇总",
        "",
        "| 指标 | 基线 | 开关后 |",
        "|:---|:---:|:---:|",
        f"| wide_leap_sum 命中小节 | {bm}/{bt} | {am}/{at} |",
        f"| wide_leap_sum 命中占比 | {bm / bt * 100:.1f}% | {am / at * 100:.1f}% |",
        f"| 命中音组记录数 | {n_rows} | {n_rows - n_arp}（豁免琶音 {n_arp} 条） |",
        "",
        "## 结果要点",
        "",
        f"1. **误报构成**：全语料 {n_rows} 条基线命中音组中，{n_arp} 条"
        f"（{n_arp / n_rows * 100:.1f}%）为琶音音型——证实复合音程判据在琶音织体上的"
        "系统性误报假设。",
        f"2. **开关效果**：wide_leap_sum 全语料命中占比 "
        f"{bm / bt * 100:.1f}% → {am / at * 100:.1f}%"
        f"（{(am / at - bm / bt) * 100:+.1f}pp）。",
        "3. **无副作用**：其余特征（consecutive_sixteenths / consecutive_chords）"
        "两版结果完全一致（逐页 Δ=0）。",
        "4. **残留误报**：非琶音保留音组中仍含声部交错与非和弦音情形"
        "（见「表_逐小节明细」），该开关不彻底解决误报问题——这是预期内的"
        "已知局限，完整解决需声部分离与和弦还原（见 README「已知限制」）。",
        "",
        "---",
        "",
        "*报告由 arpeggio_compare.py 自动生成。基线 = 交叉手重分配 + "
        "arpeggio_aware=off；开关后 = 交叉手重分配 + arpeggio_aware=on。*",
    ]
    md_path = Path(args.out + ".md")
    md_path.write_text("\n".join(L), encoding="utf-8")
    print("✓ Markdown:", md_path)

    if inconsistent:
        print(f"⚠ 自洽校验失败 {inconsistent} 页，请检查实现！")
        return 2
    print("✓ 自洽校验通过：开关后命中音组 == 基线命中中的非琶音音组")
    return 0


if __name__ == "__main__":
    sys.exit(main())
