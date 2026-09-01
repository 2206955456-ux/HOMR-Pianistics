# -*- coding: utf-8 -*-
"""报告输出：控制台摘要 + Markdown 报告 + JSON + Excel（供程序消费）。"""

import json
import time
from pathlib import Path

from scoreflow.analysis.engine import ScoreResult


def print_summary(results: list[ScoreResult]) -> None:
    for res in results:
        print(f"\n■ {res.source.name if res.source else '<memory>'}")
        for name, fres in res.feature_results.items():
            print(f"  [{fres.feature.name}] {fres.feature.description}")
            if fres.feature.describe_params():
                print(f"      参数: {fres.feature.describe_params()}")
            print(f"      命中小节: {fres.matched_measures}/{fres.total_measures}"
                  f"  占比 {fres.ratio_str}")
        if res.difficulty_detail:
            print(f"  ★ 难度加权分（《音型集》口径）: {res.difficulty_score:g}")
            for fname, d in res.difficulty_detail.items():
                two = d["two_hand_measures"]
                print(f"      {fname}: 权重{d['weight']:g} × (单手{d['hit_measures']-two}"
                      f" + 双手{two}×2) = {d['score']:g}")


def write_markdown(results: list[ScoreResult], out_path: Path, source_pdf: str = "") -> Path:
    lines = [
        "# 乐谱音群特征分析报告",
        "",
        f"- 生成时间: {time.strftime('%Y-%m-%d %H:%M:%S')}",
        f"- 源文件: {source_pdf or (results[0].source.name if results else '')}",
        f"- 分析页数: {sum(r.pages for r in results)}",
        "",
    ]
    for res in results:
        lines.append(f"## {res.source.name if res.source else '<memory>'}")
        lines.append("")
        for name, fres in res.feature_results.items():
            params = fres.feature.describe_params()
            lines += [
                f"### {fres.feature.name} — {fres.feature.description}"
                + (f"（{params}）" if params else ""),
                "",
                f"- 命中小节: **{fres.matched_measures} / {fres.total_measures}**",
                f"- 占全曲比例: **{fres.ratio_str}**",
                "",
            ]
            if fres.hits:
                lines.append("| 小节 | 声部 | 命中音群 |")
                lines.append("|---:|:---|:---|")
                for h in fres.hits:
                    lines.append(f"| {h.measure_number} | {h.part_name} | {h.notes_desc} |")
                lines.append("")
        if res.difficulty_detail:
            lines += [
                "### 难度加权（《音型集》口径）",
                "",
                f"**总分：{res.difficulty_score:g}**",
                "",
                "| 特征 | 权重 | 命中小节 | 其中双手同节 | 加权得分 |",
                "|:---|---:|---:|---:|---:|",
            ]
            for fname, d in res.difficulty_detail.items():
                lines.append(
                    f"| {fname} | {d['weight']:g} | {d['hit_measures']} "
                    f"| {d['two_hand_measures']} | {d['score']:g} |")
            lines.append("")
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text("\n".join(lines), encoding="utf-8")
    return out_path


def write_json(results: list[ScoreResult], out_path: Path, source_pdf: str = "") -> Path:
    data = {
        "generated_at": time.strftime("%Y-%m-%d %H:%M:%S"),
        "source_pdf": source_pdf,
        "scores": [],
    }
    for res in results:
        data["scores"].append({
            "musicxml": str(res.source) if res.source else None,
            "pages": res.pages,
            "difficulty_score": res.difficulty_score,
            "difficulty_detail": res.difficulty_detail,
            "features": {
                name: {
                    "description": fres.feature.description,
                    "params": fres.feature.params,
                    "matched_measures": fres.matched_measures,
                    "total_measures": fres.total_measures,
                    "ratio": round(fres.ratio, 4),
                    "hits": [
                        {
                            "measure": h.measure_number,
                            "part": h.part_name,
                            "notes": h.notes_desc,
                        }
                        for h in fres.hits
                    ],
                }
                for name, fres in res.feature_results.items()
            },
        })
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(
        json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    return out_path


def write_excel(results: list[ScoreResult], out_path: Path, source_pdf: str = "") -> Path:
    """输出 Excel 报告：Sheet「总览」= 每特征一行汇总；每个特征一张明细表。

    依赖 openpyxl（已列入 requirements.txt）。缺失时给出明确提示。
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise RuntimeError(
            "需要 openpyxl 才能输出 Excel 报告，请先安装："
            "python -m pip install openpyxl")

    HEADER_FILL = PatternFill("solid", fgColor="4472C4")
    HEADER_FONT = Font(bold=True, color="FFFFFF")
    HIT_FILL = PatternFill("solid", fgColor="FFF2CC")

    wb = Workbook()

    # ---------- Sheet 1: 总览 ----------
    ws = wb.active
    ws.title = "总览"
    headers = ["乐谱", "特征", "说明", "参数", "命中小节数", "总小节数", "占比",
               "权重", "双手同节", "加权得分"]
    ws.append(headers)
    # 多首曲子逐首追加（每首一行标题分隔）
    for res in results:
        score_name = res.source.name if res.source else "<memory>"
        for fname, fres in res.feature_results.items():
            d = res.difficulty_detail.get(fname, {})
            ws.append([
                score_name,
                fres.feature.name,
                fres.feature.description,
                fres.feature.describe_params(),
                fres.matched_measures,
                fres.total_measures,
                round(fres.ratio * 100, 1),
                d.get("weight", ""),
                d.get("two_hand_measures", ""),
                d.get("score", ""),
            ])
    # 样式：表头
    for col in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="center", vertical="center")
    # 列宽
    widths = [24, 26, 46, 18, 12, 12, 10, 8, 10, 10]
    for col, width in zip(range(1, len(widths) + 1), widths):
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.freeze_panes = "A2"
    # 占比列格式
    for r in range(2, ws.max_row + 1):
        ws.cell(row=r, column=7).number_format = "0.0\"%\""
    # 合计行（仅单曲时有意义；多曲时占比不可直接相加，不输出避免误导）
    n_scores = len(results)
    if n_scores == 1:
        ws.append([])
        total_row = ws.max_row + 1  # 合计行即将写入的行号
        data_last = total_row - 2   # 数据末行
        ws.append([
            "合计", "", "", "",
            f"=SUM(E2:E{data_last})",
            f"=SUM(F2:F{data_last})",
            f'=IF(F{data_last}=0,"",ROUND(E{total_row}/F{total_row}*100,1))',
        ])

    # ---------- Sheet 2: 难度加权汇总 ----------
    if any(res.difficulty_detail for res in results):
        wq = wb.create_sheet("难度加权")
        wq.append(["乐谱", "特征", "权重", "命中小节数",
                   "其中双手同节", "单手小节数", "加权得分"])
        for res in results:
            score_name = res.source.name if res.source else "<memory>"
            for fname, d in res.difficulty_detail.items():
                two = d["two_hand_measures"]
                wq.append([score_name, fname, d["weight"], d["hit_measures"],
                           two, d["hit_measures"] - two, d["score"]])
            if res.difficulty_detail:
                wq.append([score_name, "★总分", "", "", "", "", res.difficulty_score])
        for col in range(1, 8):
            c = wq.cell(row=1, column=col)
            c.fill = HEADER_FILL
            c.font = HEADER_FONT
            c.alignment = Alignment(horizontal="center")
        for col, width in zip(range(1, 8), [24, 26, 8, 12, 12, 12, 10]):
            wq.column_dimensions[get_column_letter(col)].width = width
        wq.freeze_panes = "A2"

    # ---------- 每个特征一张明细表 ----------
    feature_names = sorted({
        fname
        for res in results
        for fname in res.feature_results
    })
    for fname in feature_names:
        wsd = wb.create_sheet(f"明细_{fname}"[:31])  # Excel 表名最长 31 字符
        wsd.append(["乐谱", "小节", "声部", "命中音群"])
        for res in results:
            fres = res.feature_results.get(fname)
            if not fres or not fres.hits:
                continue
            score_name = res.source.name if res.source else "<memory>"
            for h in fres.hits:
                wsd.append([score_name, h.measure_number, h.part_name, h.notes_desc])
        for col in range(1, 5):
            c = wsd.cell(row=1, column=col)
            c.fill = HEADER_FILL
            c.font = HEADER_FONT
            c.alignment = Alignment(horizontal="center")
        for col, width in zip(range(1, 5), [24, 8, 14, 70]):
            wsd.column_dimensions[get_column_letter(col)].width = width
        wsd.freeze_panes = "A2"
        # 命中小节行高亮
        for r in range(2, wsd.max_row + 1):
            wsd.cell(row=r, column=2).fill = HIT_FILL

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out_path))
    return out_path
