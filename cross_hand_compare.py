# -*- coding: utf-8 -*-
"""交叉手启发式：基线 vs 重分配 论文级对比报告生成器。

扫描 output 目录下所有乐谱（每曲 pages/page_*.musicxml），对每页分别跑
基线引擎与启用了交叉手启发式的引擎，汇总生成：

  Excel「交叉手对比_论文版.xlsx」
    - 表_对比总览    ：每曲 × 每特征 的两版命中/占比/Δ（论文主表）
    - 表_按特征汇总  ：跨曲目的每特征合计对比（论文次要表）
    - 表_逐页命中明细：每曲每页每特征的两版命中
    - 表_重分配音符明细：每个被重分配的音符（曲目/页/小节/音符/方向/voice/staff）
    - 表_声部分布    ：每曲每特征每声部的命中记录数对比

  Markdown「交叉手对比_论文版.md」：可直接粘贴进论文的表格。

用法：
    python cross_hand_compare.py                      # 默认扫 output 目录
    python cross_hand_compare.py -o 输出目录           # 指定 XML 根目录
    python cross_hand_compare.py --out 保存前缀        # 输出文件前缀
"""

import argparse
import re
import sys
from collections import Counter, defaultdict
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))

from scoreflow.analysis.engine import AnalysisEngine
from scoreflow.analysis.features import build_features
from scoreflow.config import load_config


def run_both(engine_base, engine_ch, xml: Path):
    """对单个 MusicXML 分别跑基线与启发式引擎。返回 (基线ScoreResult, 启发式ScoreResult)。"""
    rb = engine_base.analyze_file(xml)
    rc = engine_ch.analyze_file(xml)
    return rb, rc


def collect_scores(xml_root: Path, pattern: str = "etude*") -> list[tuple[str, list[Path]]]:
    scores = []
    for d in sorted(xml_root.glob(f"{pattern}/pages")):
        xmls = sorted(d.glob("page_*.musicxml"))
        if xmls:
            scores.append((d.parent.name, xmls))
    return scores


def agg(results, feature_name):
    """把多页 ScoreResult 按特征累加 matched/total，hits 汇总。"""
    total = sum(r.feature_results[feature_name].total_measures for r in results)
    matched = sum(r.feature_results[feature_name].matched_measures for r in results)
    hits = []
    for r in results:
        hits.extend(r.feature_results[feature_name].hits)
    return total, matched, hits


def main():
    ap = argparse.ArgumentParser(description="交叉手启发式对比报告")
    ap.add_argument("-o", "--xml-root", default=r"D:\HOMR\output",
                    help="XML 根目录（含 <曲目>/pages/page_*.musicxml）")
    ap.add_argument("--pattern", default="etude*",
                    help="曲目目录过滤（glob），默认 etude* 只处理考级练习曲")
    ap.add_argument("--out", default=r"D:\HOMR\output\reports\交叉手对比_论文版",
                    help="输出文件前缀（.xlsx / .md 自动追加）")
    args = ap.parse_args()

    cfg = load_config(None, {})
    eng_base = AnalysisEngine(
        build_features(cfg.features), cfg.chord_mode,
        cfg.count_empty_measures, cfg.across_barlines, cross_hand_heuristic=False)
    eng_ch = AnalysisEngine(
        build_features(cfg.features), cfg.chord_mode,
        cfg.count_empty_measures, cfg.across_barlines, cross_hand_heuristic=True)
    feature_names = [f.name for f in eng_base.features]

    scores = collect_scores(Path(args.xml_root), args.pattern)
    if not scores:
        print("✗ 未找到乐谱目录（*/pages/page_*.musicxml）:", args.xml_root)
        return 1
    print(f"找到 {len(scores)} 首乐谱，特征: {feature_names}")

    # ===== 汇总数据结构 =====
    # overview[score][feat] = {base_t, base_m, ch_t, ch_m, ops, parts_base, parts_ch}
    overview: dict[str, dict[str, dict]] = defaultdict(lambda: defaultdict(dict))
    page_rows = []          # 逐页明细
    op_rows = []            # 重分配音符明细
    all_ops_count = Counter()

    for score_name, xmls in scores:
        base_results, ch_results = [], []
        for xml in xmls:
            rb, rc = run_both(eng_base, eng_ch, xml)
            base_results.append(rb)
            ch_results.append(rc)
            # 逐页明细
            for fn in feature_names:
                bt = rb.feature_results[fn].total_measures
                bm = rb.feature_results[fn].matched_measures
                ct = rc.feature_results[fn].total_measures
                cm = rc.feature_results[fn].matched_measures
                page_rows.append((score_name, xml.stem.replace("page_", ""),
                                  fn, bm, bt, cm, ct))
            # 重分配音符明细
            for op in rc.cross_hand_ops:
                v_new = "1" if op["voice"] == "5" else "5"
                s_new = "1" if op["staff"] == "2" else "2"
                op_rows.append((score_name, xml.stem.replace("page_", ""),
                                op["measure"], op["note"], op["midi"], op["kind"],
                                op["voice"], op["staff"], v_new, s_new))
                all_ops_count[op["kind"]] += 1
        # 汇总每曲每特征
        for fn in feature_names:
            bt, bm, bhits = agg(base_results, fn)
            ct, cm, chits = agg(ch_results, fn)
            parts_base = Counter(h.part_name for h in bhits)
            parts_ch = Counter(h.part_name for h in chits)
            overview[score_name][fn] = {
                "base_t": bt, "base_m": bm, "ch_t": ct, "ch_m": cm,
                "parts_base": parts_base, "parts_ch": parts_ch,
            }

    # ===== 写入 Excel =====
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    HEADER_FILL = PatternFill("solid", fgColor="4472C4")
    HEADER_FONT = Font(bold=True, color="FFFFFF")
    DIFF_FILL = PatternFill("solid", fgColor="E2EFDA")   # 有变化（绿）
    NOCHG_FILL = PatternFill("solid", fgColor="F2F2F2")  # 无变化（灰）

    wb = Workbook()

    # ---- Sheet1 对比总览 ----
    ws = wb.active
    ws.title = "表_对比总览"
    headers = ["乐谱", "小节数", "特征", "基线命中/小节", "基线占比%",
               "启发式命中/小节", "启发式占比%", "Δ占比(pp)", "重分配音符数"]
    ws.append(headers)
    for score_name, feats in overview.items():
        n_measures = max(f["base_t"] for f in feats.values())
        for fn in feature_names:
            f = feats[fn]
            b_ratio = f["base_m"] / f["base_t"] * 100 if f["base_t"] else 0.0
            c_ratio = f["ch_m"] / f["ch_t"] * 100 if f["ch_t"] else 0.0
            delta = round(c_ratio - b_ratio, 1)
            ops_n = sum(1 for o in op_rows if o[0] == score_name)  # 简化：用特征无关的总数
            ws.append([score_name, n_measures, fn,
                       f"{f['base_m']}/{f['base_t']}", round(b_ratio, 1),
                       f"{f['ch_m']}/{f['ch_t']}", round(c_ratio, 1),
                       delta, ops_n])
    for col in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="center", vertical="center")
    for r in range(2, ws.max_row + 1):
        # 占比列格式
        for col in (5, 7):
            ws.cell(row=r, column=col).number_format = "0.0\"%\""
        # Δ列：有变化绿色、无变化灰色
        dcell = ws.cell(row=r, column=8)
        if dcell.value == 0:
            dcell.fill = NOCHG_FILL
        else:
            dcell.fill = DIFF_FILL
    for col, width in zip(range(1, 10), [16, 9, 26, 14, 11, 14, 11, 11, 12]):
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.freeze_panes = "A2"

    # ---- Sheet2 按特征汇总 ----
    ws2 = wb.create_sheet("表_按特征汇总")
    ws2.append(["特征", "乐谱数", "基线命中合计", "基线小节合计", "基线占比%",
                "启发式命中合计", "启发式小节合计", "启发式占比%", "Δ占比(pp)"])
    for fn in feature_names:
        bm = sum(f[fn]["base_m"] for f in overview.values())
        bt = sum(f[fn]["base_t"] for f in overview.values())
        cm = sum(f[fn]["ch_m"] for f in overview.values())
        ct = sum(f[fn]["ch_t"] for f in overview.values())
        ws2.append([fn, len(overview), bm, bt,
                    round(bm / bt * 100, 1) if bt else 0,
                    cm, ct,
                    round(cm / ct * 100, 1) if ct else 0,
                    round((cm / ct - bm / bt) * 100, 1) if bt and ct else 0])
    for col in range(1, 10):
        c = ws2.cell(row=1, column=col)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="center")
    for r in range(2, ws2.max_row + 1):
        for col in (5, 8):
            ws2.cell(row=r, column=col).number_format = "0.0\"%\""
    for col, width in zip(range(1, 10), [26, 9, 13, 13, 11, 13, 13, 11, 11]):
        ws2.column_dimensions[get_column_letter(col)].width = width
    ws2.freeze_panes = "A2"

    # ---- Sheet3 逐页命中明细 ----
    ws3 = wb.create_sheet("表_逐页命中明细")
    ws3.append(["乐谱", "页", "特征", "基线命中/小节", "启发式命中/小节", "Δ"])
    for score_name, page_no, fn, bm, bt, cm, ct in page_rows:
        ws3.append([score_name, int(page_no), fn,
                    f"{bm}/{bt}", f"{cm}/{ct}", cm - bm])
    for col in range(1, 7):
        c = ws3.cell(row=1, column=col)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="center")
    for col, width in zip(range(1, 7), [16, 7, 26, 14, 14, 6]):
        ws3.column_dimensions[get_column_letter(col)].width = width
    ws3.freeze_panes = "A2"

    # ---- Sheet4 重分配音符明细 ----
    ws4 = wb.create_sheet("表_重分配音符明细")
    ws4.append(["乐谱", "页", "小节", "音符", "midi", "方向", "原voice", "原staff", "新voice", "新staff"])
    for row in op_rows:
        ws4.append(list(row))
    for col in range(1, 11):
        c = ws4.cell(row=1, column=col)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="center")
    for col, width in zip(range(1, 11), [16, 7, 7, 8, 7, 10, 9, 9, 9, 9]):
        ws4.column_dimensions[get_column_letter(col)].width = width
    ws4.freeze_panes = "A2"

    # ---- Sheet5 声部分布 ----
    ws5 = wb.create_sheet("表_声部分布")
    ws5.append(["乐谱", "特征", "声部", "基线命中记录数", "启发式命中记录数", "Δ"])
    for score_name, feats in overview.items():
        for fn in feature_names:
            f = feats[fn]
            all_parts = sorted(set(f["parts_base"]) | set(f["parts_ch"]))
            for p in all_parts:
                b = f["parts_base"].get(p, 0)
                c = f["parts_ch"].get(p, 0)
                ws5.append([score_name, fn, p, b, c, c - b])
    for col in range(1, 7):
        c = ws5.cell(row=1, column=col)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="center")
    for col, width in zip(range(1, 7), [16, 26, 12, 14, 14, 6]):
        ws5.column_dimensions[get_column_letter(col)].width = width
    ws5.freeze_panes = "A2"

    xlsx_path = Path(args.out + ".xlsx")
    xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(xlsx_path))
    print("✓ Excel:", xlsx_path)

    # ===== 写入 Markdown =====
    lines = [
        "# 交叉手启发式声部重分配：基线 vs 重分配对比报告",
        "",
        f"- 生成时间: {__import__('time').strftime('%Y-%m-%d %H:%M:%S')}",
        f"- XML 根目录: {args.xml_root}",
        f"- 重分配规则: staff2 上 >C5(midi72) → 右手；staff1 上 <C3(midi48) → 左手",
        f"- 重分配音符合计: {sum(all_ops_count.values())}（{dict(all_ops_count)}）",
        "",
        "## 表 1. 每首乐谱 × 每特征对比（论文主表）",
        "",
        "| 乐谱 | 小节数 | 特征 | 基线命中/小节 | 基线占比 | 启发式命中/小节 | 启发式占比 | Δ占比(pp) |",
        "|---:|---:|---|:---:|:---:|:---:|:---:|:---:|",
    ]
    for score_name, feats in overview.items():
        n_measures = max(f["base_t"] for f in feats.values())
        for fn in feature_names:
            f = feats[fn]
            b_ratio = f["base_m"] / f["base_t"] * 100 if f["base_t"] else 0.0
            c_ratio = f["ch_m"] / f["ch_t"] * 100 if f["ch_t"] else 0.0
            lines.append(
                f"| {score_name} | {n_measures} | {fn} "
                f"| {f['base_m']}/{f['base_t']} | {b_ratio:.1f}% "
                f"| {f['ch_m']}/{f['ch_t']} | {c_ratio:.1f}% "
                f"| {c_ratio - b_ratio:+.1f} |")
    lines += [
        "",
        "## 表 2. 按特征跨曲目汇总",
        "",
        "| 特征 | 基线命中/小节 | 基线占比 | 启发式命中/小节 | 启发式占比 | Δ占比(pp) |",
        "|:---|---:|:---:|---:|:---:|:---:|",
    ]
    for fn in feature_names:
        bm = sum(f[fn]["base_m"] for f in overview.values())
        bt = sum(f[fn]["base_t"] for f in overview.values())
        cm = sum(f[fn]["ch_m"] for f in overview.values())
        ct = sum(f[fn]["ch_t"] for f in overview.values())
        lines.append(
            f"| {fn} | {bm}/{bt} | {bm/bt*100:.1f}% "
            f"| {cm}/{ct} | {cm/ct*100:.1f}% "
            f"| {(cm/ct - bm/bt)*100:+.1f} |")
    lines += [
        "",
        "## 表 3. 重分配音符统计（按乐谱）",
        "",
        "| 乐谱 | LH→RH（左手跑高音区） | RH→LH（右手跑低音区） | 合计 |",
        "|:---|---:|---:|---:|",
    ]
    for score_name, feats in overview.items():
        ops_for_score = [o for o in op_rows if o[0] == score_name]
        lh = sum(1 for o in ops_for_score if o[5] == "LH->RH")
        rh = sum(1 for o in ops_for_score if o[5] == "RH->LH")
        lines.append(f"| {score_name} | {lh} | {rh} | {lh+rh} |")
    lines.append(f"| **合计** | **{all_ops_count['LH->RH']}** | **{all_ops_count['RH->LH']}** | **{sum(all_ops_count.values())}** |")
    lines += [
        "",
        "## 表 4. 声部分布（命中记录数，wide_leap_sum）",
        "",
        "| 乐谱 | 声部 | 基线 | 启发式 | Δ |",
        "|:---|:---|---:|---:|---:|",
    ]
    for score_name, feats in overview.items():
        f = feats["wide_leap_sum"]
        for p in sorted(set(f["parts_base"]) | set(f["parts_ch"])):
            b = f["parts_base"].get(p, 0)
            c = f["parts_ch"].get(p, 0)
            lines.append(f"| {score_name} | {p} | {b} | {c} | {c-b:+d} |")
    lines += [
        "",
        "---",
        "",
        "*报告由 cross_hand_compare.py 自动生成；基线=按谱表分组（HOMR 原始 voice/staff），",
        "启发式=音域阈值重分配（模拟符干朝向判断的近似）。",
        "阈值为 C3(midi48)/C5(midi72)，双手音域重叠段可能误判，详见 README「已知限制」。*",
    ]
    md_path = Path(args.out + ".md")
    md_path.write_text("\n".join(lines), encoding="utf-8")
    print("✓ Markdown:", md_path)
    return 0


if __name__ == "__main__":
    sys.exit(main())
