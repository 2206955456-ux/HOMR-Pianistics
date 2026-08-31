#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
agreement_eval —— 音群特征自动检测 vs 人工标注的一致性评估（论文 6.2 节主实验）

用法：
    # 1) 生成标注模板（随机抽样小节，3 个特征列留空待填）
    python agreement_eval.py --make-template --out 标注模板.xlsx --per-score 10 --seed 42

    # 2) 用户在 Excel 中填写（1=该小节含该特征，0=不含，?=不确定）
    # 3) 读回标注并计算 精确率/召回率/F1/Cohen's Kappa
    python agreement_eval.py --annot 标注模板_已填.xlsx --out 一致性评估.xlsx

说明：
  - 标注单位 = (曲目, 页, 小节号)：MusicXML 每页小节号重新编号，必须带页码。
  - 自动标签在评估时对每页 MusicXML 实时重算（不依赖 report.json 的 hits，避免页码歧义），
    因此若重新跑过 OMR，请重新生成模板。
  - 1=命中，0=未命中；'?'或留空视为无效样本，不计入指标。
"""

import argparse
import json
import random
import sys
import warnings
from pathlib import Path

warnings.filterwarnings("ignore")
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")

sys.path.insert(0, str(Path(__file__).resolve().parent))

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from music21 import chord, converter

from scoreflow.analysis.engine import AnalysisEngine
from scoreflow.analysis.features import build_features
from scoreflow.config import load_config

DEFAULT_REPORT = r"D:\HOMR\output\reports\report.json"
HEADER_FILL = PatternFill("solid", fgColor="D9E1F2")      # 表头蓝
FILL_ME = PatternFill("solid", fgColor="FFF2CC")           # 待填黄
OK_FILL = PatternFill("solid", fgColor="E2EFDA")           # 完成绿


# ---------------------------------------------------------------------------
def build_engine(cfg) -> AnalysisEngine:
    return AnalysisEngine(
        features=build_features(cfg.features),
        chord_mode=cfg.chord_mode,
        count_empty_measures=cfg.count_empty_measures,
        across_barlines=cfg.across_barlines,
    )


def page_info(engine, page_xml: Path):
    """解析单页 MusicXML，返回 (自动命中集合 dict, 参与统计的小节号列表)。"""
    score = converter.parse(str(page_xml))
    result = engine.analyze_score(score, source=page_xml)
    matched = {
        name: {h.measure_number for h in fres.hits}
        for name, fres in result.feature_results.items()
    }
    # 与引擎口径一致：任一特征（压平流或和弦原始流）有音符的小节计入分母
    needs_raw = any(getattr(f, "needs_chords", False) for f in engine.features)
    counted: set[int] = set()
    parts = score.parts if score.parts else [score]
    for part in parts:
        for m in part.getElementsByClass("Measure"):
            mnum = int(m.number)
            flat = engine._collect_notes(m, keep_chords=False)
            raw = engine._collect_notes(m, keep_chords=True) if needs_raw else []
            if flat or raw:
                counted.add(mnum)
    return matched, sorted(counted)


def load_scores_meta(report_json: Path, output_root: Path):
    """从 report.json 提取各曲目的页面 MusicXML 列表。"""
    d = json.loads(report_json.read_text(encoding="utf-8"))
    metas = []
    for s in d["scores"]:
        stem = Path(s["musicxml"]).stem
        pages_dir = output_root / stem / "pages"
        xmls = sorted(pages_dir.glob("page_*.musicxml"))
        if not xmls:
            print(f"  ! {stem}: 无页面 MusicXML，跳过")
            continue
        metas.append({
            "stem": stem,
            "xmls": xmls,
            "pngs": {p.stem: p for p in pages_dir.glob("page_*.png")},
        })
    return metas


# ---------------------------------------------------------------------------
def make_template(args):
    cfg = load_config(args.config)
    output_root = Path(args.output_root)
    engine = build_engine(cfg)
    feats = [f.name for f in engine.features]
    metas = load_scores_meta(Path(args.report), output_root)
    rng = random.Random(args.seed)

    samples = []   # dict(score, page, measure, img)
    for meta in metas:
        pool = []  # (page_idx, page_no, measure)
        for page_idx, xml in enumerate(meta["xmls"], 1):
            _, counted = page_info(engine, xml)
            pool.extend((page_idx, page_idx, m) for m in counted)
        if not pool:
            print(f"  ! {meta['stem']}: 无含音符小节")
            continue
        n = min(args.per_score, len(pool))
        for page_idx, page_no, m in rng.sample(pool, n):
            png = meta["pngs"].get(f"page_{page_no:03d}")
            samples.append({
                "score": meta["stem"],
                "page": page_no,
                "measure": m,
                "img": str(png) if png else "",
            })
    samples.sort(key=lambda s: (s["score"], s["page"], s["measure"]))

    wb = Workbook()
    ws = wb.active
    ws.title = "标注表"
    # 两行表头：第一行特征名，第二行定义/参数
    ws.cell(1, 1, "序号"); ws.cell(1, 2, "曲目"); ws.cell(1, 3, "页")
    ws.cell(1, 4, "小节号"); ws.cell(1, 5, "查看图片(PNG)")
    ws.cell(2, 1, ""); ws.cell(2, 2, ""); ws.cell(2, 3, "")
    ws.cell(2, 4, ""); ws.cell(2, 5, "双击打开该页乐谱图，对照标注")
    for i, f in enumerate(feats, start=6):
        ws.cell(1, i, f)
        cls = next(x for x in engine.features if x.name == f)
        ws.cell(2, i, cls.description)
    for c in range(1, 6 + len(feats)):
        ws.cell(1, c).fill = HEADER_FILL
        ws.cell(2, c).fill = HEADER_FILL
        ws.cell(1, c).font = Font(bold=True)
        ws.cell(2, c).font = Font(size=9, italic=True)
        ws.cell(1, c).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.cell(2, c).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for idx, s in enumerate(samples, 1):
        r = idx + 2
        ws.cell(r, 1, idx)
        ws.cell(r, 2, s["score"])
        ws.cell(r, 3, s["page"])
        ws.cell(r, 4, s["measure"])
        ws.cell(r, 5, s["img"])
        for i in range(len(feats)):
            cell = ws.cell(r, 6 + i)
            cell.fill = FILL_ME
        ws.cell(r, 6 + len(feats) + 1, "")  # 备注列
    ws.cell(1, 6 + len(feats) + 1, "备注")
    ws.cell(1, 6 + len(feats) + 1).fill = HEADER_FILL

    dv = DataValidation(type="list", formula1='"0,1,?"', allow_blank=True)
    dv.error = "只能填 0、1 或 ?（?=不确定，不计入指标）"
    dv.errorTitle = "非法输入"
    first = get_column_letter(6)
    last = get_column_letter(5 + len(feats))
    dv.add(f"{first}3:{last}{len(samples) + 2}")
    ws.add_data_validation(dv)

    ws.freeze_panes = "C3"
    widths = [6, 12, 6, 8, 46] + [12] * len(feats) + [12]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # 说明 sheet
    sh = wb.create_sheet("标注说明")
    lines = [
        "【标注规则】",
        "1. 每一行对应 (曲目, 页, 小节号) 的一个小节；'查看图片'列双击可打开该页乐谱渲染图（PDF 第 N 页）。",
        "2. 三个特征列填：1 = 该小节内存在至少一处该特征；0 = 不存在；? = 不确定（不计入指标）。",
        "3. 判断的是'小节内是否有该技术型音群'，与声部无关——任一声部出现即算 1。",
        "4. 请在真实乐谱上独立判断，先不要参考系统报告，避免标注偏差。",
        "",
        "【特征定义】",
    ]
    for f in engine.features:
        lines.append(f"  - {f.name}: {f.description}")
    sh.column_dimensions["A"].width = 110
    for i, t in enumerate(lines, 1):
        sh.cell(i, 1, t).alignment = Alignment(wrap_text=True, vertical="top")
    sh.cell(1, 1).font = Font(bold=True, size=12)

    out = Path(args.out)
    wb.save(out)
    print(f"标注模板已生成: {out}")
    print(f"  抽样: {len(samples)} 小节（每首 ≤{args.per_score}，seed={args.seed}）")
    print(f"  特征: {', '.join(feats)}")
    print("请在 Excel 中逐行填写 1/0/?，保存后运行:")
    print(f"  python agreement_eval.py --annot \"{out.name}\"")


# ---------------------------------------------------------------------------
def cohen_kappa(tp, fp, fn, tn):
    n = tp + fp + fn + tn
    if n == 0:
        return float("nan")
    p0 = (tp + tn) / n
    pe = ((tp + fp) * (tp + fn) + (fp + tn) * (fn + tn)) / (n * n)
    return (p0 - pe) / (1 - pe) if pe != 1 else float("nan")


def kappa_level(k):
    if k != k:  # nan
        return "-"
    if k < 0.2:
        return "极低"
    if k < 0.4:
        return "一般"
    if k < 0.6:
        return "中等"
    if k < 0.8:
        return "高"
    return "极高"


def evaluate(args):
    cfg = load_config(args.config)
    output_root = Path(args.output_root)
    engine = build_engine(cfg)
    feats = [f.name for f in engine.features]

    wb_in = load_workbook(args.annot, data_only=True)
    ws = wb_in["标注表"] if "标注表" in wb_in.sheetnames else wb_in.active
    feats_col = {ws.cell(1, c).value: c for c in range(6, 6 + len(feats))}
    feats_col = {k: v for k, v in feats_col.items() if k}
    rows = []
    for r in range(3, ws.max_row + 1):
        score = ws.cell(r, 2).value
        page = ws.cell(r, 3).value
        measure = ws.cell(r, 4).value
        if score is None or page is None or measure is None:
            continue
        labels = {}
        for f, c in feats_col.items():
            v = ws.cell(r, c).value
            labels[f] = v
        rows.append({"row": r, "score": str(score).strip(),
                     "page": int(page), "measure": int(measure), "labels": labels})

    # 逐页重算自动标签（缓存）
    auto_cache: dict[tuple[str, int], dict[str, set]] = {}
    for s in rows:
        key = (s["score"], s["page"])
        if key in auto_cache:
            continue
        pages_dir = output_root / s["score"] / "pages"
        xml = pages_dir / f"page_{s['page']:03d}.musicxml"
        if xml.exists():
            matched, _ = page_info(engine, xml)
            auto_cache[key] = matched
        else:
            auto_cache[key] = None

    # 统计
    stats = {f: {"tp": 0, "fp": 0, "fn": 0, "tn": 0, "n": 0}
             for f in feats}
    per_score = {}
    detail = []
    for s in rows:
        auto = auto_cache.get((s["score"], s["page"]))
        for f in feats:
            lab = s["labels"].get(f)
            if lab is None or str(lab).strip() == "" or str(lab).strip() == "?":
                continue
            lab = int(lab)
            if auto is None:
                print(f"  ! 缺少页面 MusicXML，跳过: {s['score']} 页 {s['page']}")
                continue
            auto_lab = 1 if s["measure"] in auto[f] else 0
            agree = (lab == auto_lab)
            detail.append({**s, "feature": f, "auto": auto_lab, "human": lab,
                           "agree": agree})
            k = (s["score"], f)
            ps = per_score.setdefault(k, {"tp": 0, "fp": 0, "fn": 0, "tn": 0})
            if lab == 1 and auto_lab == 1:
                stats[f]["tp"] += 1; ps["tp"] += 1
            elif lab == 1 and auto_lab == 0:
                stats[f]["fp"] += 1; ps["fp"] += 1
            elif lab == 0 and auto_lab == 1:
                stats[f]["fn"] += 1; ps["fn"] += 1
            else:
                stats[f]["tn"] += 1; ps["tn"] += 1

    # 输出 Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "指标总览"
    headers = ["特征", "TP", "FP", "FN", "TN", "样本数",
               "精确率P", "召回率R", "F1", "Cohen's Kappa", "一致性水平"]
    ws.append(headers)
    for f in feats:
        st = stats[f]
        p = st["tp"] / (st["tp"] + st["fp"]) if st["tp"] + st["fp"] else float("nan")
        r_ = st["tp"] / (st["tp"] + st["fn"]) if st["tp"] + st["fn"] else float("nan")
        f1 = 2 * p * r_ / (p + r_) if p + r_ else float("nan")
        k = cohen_kappa(st["tp"], st["fp"], st["fn"], st["tn"])
        n = st["tp"] + st["fp"] + st["fn"] + st["tn"]
        ws.append([f, st["tp"], st["fp"], st["fn"], st["tn"], n,
                   round(p, 3), round(r_, 3), round(f1, 3),
                   round(k, 3), kappa_level(k)])

    ws2 = wb.create_sheet("分曲明细")
    ws2.append(["曲目", "特征", "TP", "FP", "FN", "TN", "样本数",
                "精确率P", "召回率R", "F1", "Kappa"])
    for (score, f), st in sorted(per_score.items()):
        p = st["tp"] / (st["tp"] + st["fp"]) if st["tp"] + st["fp"] else float("nan")
        r_ = st["tp"] / (st["tp"] + st["fn"]) if st["tp"] + st["fn"] else float("nan")
        f1 = 2 * p * r_ / (p + r_) if p + r_ else float("nan")
        k = cohen_kappa(st["tp"], st["fp"], st["fn"], st["tn"])
        n = st["tp"] + st["fp"] + st["fn"] + st["tn"]
        ws2.append([score, f, st["tp"], st["fp"], st["fn"], st["tn"], n,
                    round(p, 3), round(r_, 3), round(f1, 3), round(k, 3)])

    ws3 = wb.create_sheet("样本明细")
    ws3.append(["曲目", "页", "小节号", "特征", "自动", "人工", "是否一致"])
    for d in detail:
        ws3.append([d["score"], d["page"], d["measure"], d["feature"],
                    d["auto"], d["human"], "✓" if d["agree"] else "✗"])

    out = Path(args.out)
    wb.save(out)

    # 控制台摘要
    print(f"\n一致性评估已生成: {out}\n")
    print(f"{'特征':<24}{'P':>7}{'R':>7}{'F1':>7}{'Kappa':>8}{'水平':>6}{'N':>6}")
    for f in feats:
        st = stats[f]
        p = st["tp"] / (st["tp"] + st["fp"]) if st["tp"] + st["fp"] else float("nan")
        r_ = st["tp"] / (st["tp"] + st["fn"]) if st["tp"] + st["fn"] else float("nan")
        f1 = 2 * p * r_ / (p + r_) if p + r_ else float("nan")
        k = cohen_kappa(st["tp"], st["fp"], st["fn"], st["tn"])
        n = st["tp"] + st["fp"] + st["fn"] + st["tn"]
        print(f"{f:<24}{p:>7.3f}{r_:>7.3f}{f1:>7.3f}{k:>8.3f}{kappa_level(k):>6}{n:>6}")


# ---------------------------------------------------------------------------
def main():
    ap = argparse.ArgumentParser(prog="agreement_eval")
    ap.add_argument("--make-template", action="store_true", help="生成标注模板")
    ap.add_argument("--annot", help="已填写的标注模板路径")
    ap.add_argument("--out", help="输出文件路径")
    ap.add_argument("--report", default=DEFAULT_REPORT, help="report.json 路径")
    ap.add_argument("--output-root", default=None, help="输出根目录（默认取 config.yaml）")
    ap.add_argument("--config", default=None, help="config.yaml 路径")
    ap.add_argument("--per-score", type=int, default=10, help="每首抽样小节数")
    ap.add_argument("--seed", type=int, default=42, help="抽样随机种子")
    args = ap.parse_args()

    cfg = load_config(args.config)
    args.output_root = args.output_root or str(cfg.output_dir)

    if args.make_template:
        make_template(args)
    elif args.annot:
        evaluate(args)
    else:
        ap.print_help()


if __name__ == "__main__":
    main()
