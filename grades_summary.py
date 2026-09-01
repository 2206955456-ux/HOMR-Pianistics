# -*- coding: utf-8 -*-
"""1-10 级考级练习曲分级分析汇总报告（音型集口径 + 难度加权）。

读取 output_grades/etude01..10/reports/report.json，汇总为：
  - 分级总览：每首 × 6 特征（命中/小节/占比）+ 难度加权总分
  - 难度加权明细：分特征权重、单手/双手小节、得分
  - 等级趋势：难度分与考级等级的关系（论文核心素材）
输出 Excel + Markdown 到 output_grades/reports/。

用法: python grades_summary.py [--root D:/HOMR/output_grades]
"""

import argparse
import json
import sys
from pathlib import Path

FEATURE_ORDER = [
    "wide_leap_sum", "contracted_grip", "consecutive_sixteenths",
    "consecutive_eighths", "consecutive_chords", "consecutive_intervals",
]
FEATURE_CN = {
    "wide_leap_sum": "伸张把位",
    "contracted_grip": "收缩把位",
    "consecutive_sixteenths": "连续十六分跑动",
    "consecutive_eighths": "连续八分跑动",
    "consecutive_chords": "连续和弦支撑",
    "consecutive_intervals": "连续音程支撑",
}
WEIGHTS = {
    "wide_leap_sum": 2, "contracted_grip": 1, "consecutive_sixteenths": 2,
    "consecutive_eighths": 1, "consecutive_chords": 2, "consecutive_intervals": 1,
}


def load_piece(root: Path, idx: int) -> dict | None:
    tag = f"etude{idx:02d}"
    p = root / tag / "reports" / "report.json"
    if not p.exists():
        return None
    d = json.loads(p.read_text(encoding="utf-8"))
    s = d["scores"][0]
    out = {"tag": tag, "grade": idx, "difficulty": s.get("difficulty_score", 0.0),
           "detail": s.get("difficulty_detail", {}), "features": {}}
    for fk, fv in s["features"].items():
        out["features"][fk] = {
            "matched": fv["matched_measures"], "total": fv["total_measures"],
            "ratio": fv["ratio"],
        }
    return out


def spearman(xs, ys):
    """Spearman 等级相关（无第三方依赖的手写实现）。"""
    def rank(v):
        order = sorted(range(len(v)), key=lambda i: v[i])
        r = [0.0] * len(v)
        i = 0
        while i < len(order):
            j = i
            while j + 1 < len(order) and v[order[j + 1]] == v[order[i]]:
                j += 1
            avg = (i + j) / 2 + 1
            for k in range(i, j + 1):
                r[order[k]] = avg
            i = j + 1
        return r
    rx, ry = rank(xs), rank(ys)
    n = len(xs)
    mx, my = sum(rx) / n, sum(ry) / n
    num = sum((a - mx) * (b - my) for a, b in zip(rx, ry))
    den = (sum((a - mx) ** 2 for a in rx) * sum((b - my) ** 2 for b in ry)) ** 0.5
    return num / den if den else 0.0


def main():
    sys.stdout.reconfigure(encoding="utf-8")
    ap = argparse.ArgumentParser()
    ap.add_argument("--root", default="D:/HOMR/output_grades")
    args = ap.parse_args()
    root = Path(args.root)

    pieces = []
    for i in range(1, 11):
        p = load_piece(root, i)
        if p:
            pieces.append(p)
        else:
            print(f"  ⚠ 缺少 etude{i:02d} 的报告，跳过")
    if not pieces:
        print("没有任何报告可汇总"); return

    out_dir = root / "reports"
    out_dir.mkdir(parents=True, exist_ok=True)

    # ---------- Markdown ----------
    L = ["# 考级教程 1–10 级练习曲 · 音群特征分级分析汇总", "",
         f"- 生成时间: {__import__('time').strftime('%Y-%m-%d %H:%M:%S')}",
         "- 口径: 《音型（音群）集》六类特征 + 难度加权（双手同节 ×2）",
         "- 配置: 交叉手重分配 + 和弦最高音修正（chord top）+ 单调方向条件",
         f"- 语料: {len(pieces)} 首（{'、'.join(p['tag'] for p in pieces)}）", ""]

    L.append("## 表 1 分级总览（六类特征命中小节 / 总小节）")
    L.append("")
    head = "| 曲目 | 等级 | 小节 | " + " | ".join(FEATURE_CN[f] for f in FEATURE_ORDER) + " | 难度加权总分 |"
    L.append(head)
    L.append("|---|---:|---:|" + "---:|" * len(FEATURE_ORDER) + "---:|")
    for p in pieces:
        tot = p["features"].get("wide_leap_sum", {}).get("total", 0)
        cells = []
        for f in FEATURE_ORDER:
            fv = p["features"].get(f, {})
            cells.append(f"{fv.get('matched', 0)}/{fv.get('total', 0)}")
        L.append(f"| {p['tag']} | {p['grade']} | {tot} | " + " | ".join(cells)
                 + f" | **{p['difficulty']:g}** |")
    L.append("")

    L.append("## 表 2 难度加权明细")
    L.append("")
    L.append("| 曲目 | " + " | ".join(f"{FEATURE_CN[f]}(w={WEIGHTS[f]})" for f in FEATURE_ORDER) + " | 总分 |")
    L.append("|---|" + "---:|" * (len(FEATURE_ORDER) + 1))
    for p in pieces:
        cells = []
        for f in FEATURE_ORDER:
            d = p["detail"].get(f, {})
            if d:
                two = d.get("two_hand_measures", 0)
                hit = d.get("hit_measures", 0)
                mark = f"{d['score']:g}" + (f"（双手{two}节）" if two else "")
                cells.append(mark)
            else:
                cells.append("0")
        L.append(f"| {p['tag']} | " + " | ".join(cells) + f" | **{p['difficulty']:g}** |")
    L.append("")

    # 等级趋势
    grades = [p["grade"] for p in pieces]
    scores = [p["difficulty"] for p in pieces]
    if len(pieces) >= 3:
        rho = spearman(grades, scores)
        L.append("## 等级—难度趋势")
        L.append("")
        L.append("难度加权总分按等级排序：")
        L.append("")
        L.append("| 等级 | " + " | ".join(str(g) for g in grades) + " |")
        L.append("|---|" + "---:|" * len(grades))
        L.append("| 难度分 | " + " | ".join(f"{s:g}" for s in scores) + " |")
        L.append("")
        L.append(f"Spearman 等级相关系数 ρ = **{rho:.3f}**"
                 f"（正值表示难度分随考级等级递增）。")
        L.append("")

    md_path = out_dir / "分级汇总_1-10级.md"
    md_path.write_text("\n".join(L), encoding="utf-8")
    print(f"✓ Markdown: {md_path}")

    # ---------- Excel ----------
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    HEADER_FILL = PatternFill("solid", fgColor="4472C4")
    HEADER_FONT = Font(bold=True, color="FFFFFF")
    TOTAL_FILL = PatternFill("solid", fgColor="FFF2CC")

    wb = Workbook()

    ws = wb.active
    ws.title = "分级总览"
    headers = (["曲目", "考级等级", "总小节"]
               + [f"{FEATURE_CN[f]}\n(权重{WEIGHTS[f]})" for f in FEATURE_ORDER]
               + ["难度加权总分", "难度分/小节"])
    ws.append(headers)
    for p in pieces:
        tot = p["features"].get("wide_leap_sum", {}).get("total", 0)
        row = [p["tag"], p["grade"], tot]
        for f in FEATURE_ORDER:
            fv = p["features"].get(f, {})
            row.append(f"{fv.get('matched', 0)}/{fv.get('total', 0)}")
        row.append(p["difficulty"])
        row.append(round(p["difficulty"] / tot, 3) if tot else 0)
        ws.append(row)
    for col in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col)
        c.fill = HEADER_FILL; c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for col, w in zip(range(1, len(headers) + 1),
                      [10, 9, 8] + [12] * len(FEATURE_ORDER) + [13, 11]):
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = "D2"
    for r in range(2, ws.max_row + 1):
        ws.cell(row=r, column=len(headers) - 1).fill = TOTAL_FILL

    # 难度加权明细
    ws2 = wb.create_sheet("难度加权明细")
    headers2 = (["曲目", "等级"]
                + [f"{FEATURE_CN[f]}\n权重{WEIGHTS[f]}" for f in FEATURE_ORDER] + ["总分"])
    ws2.append(headers2)
    for p in pieces:
        row = [p["tag"], p["grade"]]
        for f in FEATURE_ORDER:
            d = p["detail"].get(f, {})
            row.append(d.get("score", 0))
        row.append(p["difficulty"])
        ws2.append(row)
    # 单手/双手拆分明细
    ws2.append([])
    ws2.append(["— 单手/双手命中小节数（双手同节权重×2）—"])
    for p in pieces:
        row = [f"{p['tag']} 单手", p["grade"]]
        for f in FEATURE_ORDER:
            d = p["detail"].get(f, {})
            row.append(d.get("hit_measures", 0) - d.get("two_hand_measures", 0))
        ws2.append(row)
        row = [f"{p['tag']} 双手", ""]
        for f in FEATURE_ORDER:
            d = p["detail"].get(f, {})
            row.append(d.get("two_hand_measures", 0))
        ws2.append(row)
    for col in range(1, len(headers2) + 1):
        c = ws2.cell(row=1, column=col)
        c.fill = HEADER_FILL; c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for col, w in zip(range(1, len(headers2) + 1), [14, 7] + [13] * len(FEATURE_ORDER) + [10]):
        ws2.column_dimensions[get_column_letter(col)].width = w
    ws2.freeze_panes = "C2"

    # 等级趋势
    if len(pieces) >= 3:
        ws3 = wb.create_sheet("等级趋势")
        ws3.append(["等级", "曲目", "难度加权总分", "难度分/小节"])
        for p in pieces:
            tot = p["features"].get("wide_leap_sum", {}).get("total", 0)
            ws3.append([p["grade"], p["tag"], p["difficulty"],
                        round(p["difficulty"] / tot, 3) if tot else 0])
        ws3.append([])
        ws3.append(["Spearman ρ（等级 vs 总分）", spearman(grades, scores)])
        for col in range(1, 5):
            c = ws3.cell(row=1, column=col)
            c.fill = HEADER_FILL; c.font = HEADER_FONT
            c.alignment = Alignment(horizontal="center")
        for col, w in zip(range(1, 5), [10, 10, 14, 13]):
            ws3.column_dimensions[get_column_letter(col)].width = w

    xlsx_path = out_dir / "分级汇总_1-10级.xlsx"
    wb.save(str(xlsx_path))
    print(f"✓ Excel: {xlsx_path}")

    # 控制台摘要
    print("\n===== 分级难度加权总分 =====")
    for p in pieces:
        print(f"  {p['tag']} ({p['grade']}级): {p['difficulty']:g}")
    if len(pieces) >= 3:
        print(f"Spearman ρ = {spearman(grades, scores):.3f}")


if __name__ == "__main__":
    main()
